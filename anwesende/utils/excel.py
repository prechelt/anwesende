import collections
import typing as tg

import openpyxl

Columnsdict = tg.Mapping[str, tg.List]


def read_excel_as_columnsdict(filename: str) -> Columnsdict:
    """
    Return raw data from Excel's active sheet:
    strings will be stripped of leading/trailing whitespace;
    everything else will be converted to string.
    First row is treated as column headers; column order is kept.
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    result = collections.OrderedDict()
    for col in sheet.iter_cols(values_only=True):
        colname = col[0]
        assert isinstance(colname, str), f"type(colname) = {type(colname)}"
        result[colname] = [_cleansed(cell) for cell in col[1:]]
    return result

    
def _cleansed(cell):
    if isinstance(cell, str): 
        return cell.strip()
    else: 
        return str(cell or "")  # None or int or what-have-you


RowsListsType = tg.Mapping[str, tg.List[tg.Optional[tg.NamedTuple]]]  # sheetname -> sheetcontents


def write_excel_from_rowslists(filename: str, rowslists: RowsListsType,
                               indexcolumn=False) -> None:
    workbook = openpyxl.Workbook()
    for sheetname, rows in rowslists.items():
        sheet = workbook.create_sheet(sheetname)
        indexdigits = len(str(len(rows))) if indexcolumn else 0
        if len(rows) > 0:
            _write_column_headings(sheet, rows[0], 1, indexdigits)
        for rownum, row in enumerate(rows, start=2):
            _write_row(sheet, row, rownum, indexdigits)
    del workbook["Sheet"]  # get rid of initial default sheet
    workbook.save(filename)

 
def _write_column_headings(sheet, tupl: tg.Optional[tg.NamedTuple], 
                           rownum: int, indexdigits: int):
    # use the tuple's element names as headings
    assert tupl  # None does not occur here
    font = openpyxl.styles.Font(bold=True)
    if indexdigits:
        sheet.cell(column=1, row=1, value="index")
        sheet.cell(column=1, row=1).font = font
    for colnum, colname in enumerate(tupl._fields, start=1 + (indexdigits != 0)):
        sheet.cell(column=colnum, row=rownum, value=colname)
        sheet.cell(column=colnum, row=rownum).font = font


def _write_row(sheet, tupl: tg.Optional[tg.NamedTuple], 
               rownum: int, indexdigits: tg.Optional[int]):
    if indexdigits:
        index = str(rownum - 1).zfill(indexdigits)
        sheet.cell(column=1, row=rownum, value=index)
    if tupl is None: 
        return  # nothing else to do
    for colnum, value in enumerate(tupl, start=1 + (indexdigits != 0)):
        sheet.cell(column=colnum, row=rownum, value=value)
